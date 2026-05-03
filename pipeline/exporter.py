"""
pipeline/exporter.py
--------------------
List[MCQ valid dict] → Kahoot .xlsx / Quizizz .csv

Trả về BytesIO buffer để dùng trong Streamlit st.download_button.
"""

import io
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ── Answer index mapping ───────────────────────────────────────────
_ANSWER_TO_IDX = {"A": 1, "B": 2, "C": 3, "D": 4}
MAX_QUESTION_CHARS = 120
MAX_OPTION_CHARS = 75


def _prepare_mcqs(mcqs: list[dict]) -> list[dict]:
    """Validate and sanitize MCQs before platform export."""
    if not mcqs:
        raise ValueError("Khong co MCQ nao de export")

    cleaned = []
    for idx, mcq in enumerate(mcqs, 1):
        answer = str(mcq.get("answer", "A")).strip().upper()
        if answer not in _ANSWER_TO_IDX:
            raise ValueError(f"MCQ #{idx} co answer khong hop le: {answer}")

        item = dict(mcq)
        item["answer"] = answer
        item["question"] = _sanitize_text(item.get("question", ""), MAX_QUESTION_CHARS)
        for opt in ["A", "B", "C", "D"]:
            item[opt] = _sanitize_text(item.get(opt, ""), MAX_OPTION_CHARS)
            if not item[opt]:
                raise ValueError(f"MCQ #{idx} thieu option {opt}")
        if not item["question"]:
            raise ValueError(f"MCQ #{idx} thieu question")
        cleaned.append(item)
    return cleaned


def _sanitize_text(value, max_chars: int) -> str:
    text = str(value or "")
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", text)
    text = re.sub(r"\s*\n+\s*", " ", text)
    text = " ".join(text.split())
    return text[:max_chars].rstrip()


# ── Kahoot Excel ───────────────────────────────────────────────────

def export_kahoot(mcqs: list[dict]) -> bytes:
    """
    Tạo Kahoot-compatible Excel file.

    Kahoot format:
      Question | Answer 1 | Answer 2 | Answer 3 | Answer 4 | Time Limit | Correct Answer
    Correct Answer: số 1-4 tương ứng với Answer 1-4

    Returns: bytes (dùng trực tiếp với st.download_button)
    """
    mcqs = _prepare_mcqs(mcqs)
    wb = Workbook()
    ws = wb.active
    ws.title = "Kahoot MCQ"

    # Header style
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = [
        "Question", "Answer 1", "Answer 2", "Answer 3", "Answer 4",
        "Time Limit (sec)", "Correct Answer"
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, mcq in enumerate(mcqs, 2):
        correct_num = _ANSWER_TO_IDX.get(mcq.get("answer", "A"), 1)
        ws.cell(row=row_idx, column=1, value=mcq.get("question", ""))
        ws.cell(row=row_idx, column=2, value=mcq.get("A", ""))
        ws.cell(row=row_idx, column=3, value=mcq.get("B", ""))
        ws.cell(row=row_idx, column=4, value=mcq.get("C", ""))
        ws.cell(row=row_idx, column=5, value=mcq.get("D", ""))
        ws.cell(row=row_idx, column=6, value=30)          # 30 seconds default
        ws.cell(row=row_idx, column=7, value=correct_num)

        # Alternate row color
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="F4F6F8")
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = fill

    # Column widths
    col_widths = [60, 30, 30, 30, 30, 18, 16]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Quizizz CSV ────────────────────────────────────────────────────

def export_quizizz(mcqs: list[dict]) -> bytes:
    """
    Tạo Quizizz-compatible CSV file.

    Quizizz format:
      Question Text | Option 1 | Option 2 | Option 3 | Option 4 |
      Correct Answer | Time in seconds | Image Link (optional)

    Correct Answer: text của đáp án đúng (không phải số thứ tự)

    Returns: bytes UTF-8 với BOM (Excel đọc được tiếng Việt)
    """
    mcqs = _prepare_mcqs(mcqs)
    rows = []
    for mcq in mcqs:
        correct_letter = mcq.get("answer", "A")
        correct_text   = mcq.get(correct_letter, "")
        rows.append({
            "Question Text":    mcq.get("question", ""),
            "Option 1":         mcq.get("A", ""),
            "Option 2":         mcq.get("B", ""),
            "Option 3":         mcq.get("C", ""),
            "Option 4":         mcq.get("D", ""),
            "Correct Answer":   correct_text,
            "Time in seconds":  30,
            "Image Link":       "",
        })

    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    # UTF-8 BOM để Excel hiển thị đúng ký tự
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    return buf.getvalue()


# ── CLI test ───────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys, json

    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    # Test với sample data
    sample_mcqs = [
        {
            "question": "What does the pipe symbol | represent in regular expressions?",
            "A": "Concatenation of two patterns",
            "B": "Disjunction — matches either left or right pattern",
            "C": "Negation of a character class",
            "D": "Repetition of the previous character",
            "answer": "B",
            "explanation": "The pipe | is used for disjunction.",
            "bloom_level": "understand",
            "difficulty": "easy",
            "source_chunk": "chunk_001",
        },
        {
            "question": "Which algorithm reduces words to their root by applying rewrite rules in cascade?",
            "A": "BPE Tokenizer",
            "B": "Word2Vec",
            "C": "Porter Stemmer",
            "D": "BERT Tokenizer",
            "answer": "C",
            "explanation": "Porter Stemmer applies a series of rewrite rules (e.g., ATIONAL->ATE) in cascade.",
            "bloom_level": "remember",
            "difficulty": "easy",
            "source_chunk": "chunk_002",
        },
    ]

    kahoot_bytes = export_kahoot(sample_mcqs)
    with open("test_kahoot.xlsx", "wb") as f:
        f.write(kahoot_bytes)
    print(f"Kahoot export: test_kahoot.xlsx ({len(kahoot_bytes)} bytes)")

    quizizz_bytes = export_quizizz(sample_mcqs)
    with open("test_quizizz.csv", "wb") as f:
        f.write(quizizz_bytes)
    print(f"Quizizz export: test_quizizz.csv ({len(quizizz_bytes)} bytes)")

    print("Done. Mo file trong Excel de verify.")
